
import openpyxl as xl
from os.path import exists, basename
from cookbook import utils

class Workbook(xl.Workbook):


# +-----------------+
# | Constructor |
# +-----------------+-------------------------------------------------------------

  def __init__(self, file_path, curr_template):
    self.file_path = file_path
    self.curr_template = curr_template

    # Create new or load existing workbook
    if not exists(file_path):
      print(f"{file_path} not found.")
      print(f"Creating {file_path} workbook")
      super().__init__()
    else:
      print(f"Updating {file_path}")
      loaded_book = xl.load_workbook(file_path)
      self.__dict__.update(loaded_book.__dict__)
    
    # Set active sheet to current month
    curr_ws_name = utils.curr_month() + ' ' + str(utils.curr_year())
    if curr_ws_name not in self.sheetnames:
      print(f"Creating Sheet for {curr_ws_name}")
      new_sheet = self.create_sheet(curr_ws_name, 1)
      template = self._load_template(self.curr_template)
      self.copy_template(new_sheet, template)
      self.update_value('{template}', basename(curr_template), new_sheet, template)
    self.active = self[curr_ws_name]


# +-------------+
# | Private Helpers |
# +-------------+-------------------------------------------------------------


  def _is_value(self, str):
    if str is not None and len(str) > 1 and str[0] == '{' and str[-1] == '}':
      return True
    else:
      return False

  def _get_template_name(self, sheet):
    cell = 'A1'
    return sheet[cell].value

  def _get_template_path(self, sheet):
    return f"./templates/{self._get_template_name(sheet)}"

  def _load_template(self, template_path):
    try: 
      template = xl.load_workbook(template_path)
      return template.active
    except:
      print(f"Error: No template found at {template_path}")
      exit(1)

  def _get_template_value_locations(self, template):
    value_locations = {}
    row_count = 0
    for row in template.values: # get the row
      row_count += 1
      col_count = 65
      for entry in row: # go through each entry (column) in the row
        if self._is_value(entry):
          cell = f"{chr(col_count)}{row_count}"
          value_locations.update({cell: entry})
        col_count += 1
    return value_locations
  
  def _load_sheet(self, sheet_name):
    if type(sheet_name) == str:
      try:
        sheet = self[sheet_name]
      except:
        print("Sheet Name not found")
    elif sheet_name == 0:
      sheet = self.active
    else:
      print("Enter valid sheet_name in save_sheet_values()")
      exit(1)
    return sheet

  
  



# +-------------+
# | Attributes |
# +-------------+-------------------------------------------------------------
  def copy_template(self, sheet, template):
    row_count = 0
    for row in template.values: # get the row
      row_count += 1
      col_count = 65
      for entry in row: # go through each entry (column) in the row
        cell = f"{chr(col_count)}{row_count}"
        sheet[cell] = entry
        col_count += 1

  def sheet_values(self, sheet_name):
    sheet = self._load_sheet(sheet_name)

    # Get the location of each value from the template
    template = self._load_template(self._get_template_path(sheet))

    value_locations = self._get_template_value_locations(template)
    values = {}
    row_count = 0
    for row in sheet.values: # get the row
      row_count += 1
      col_count = 65
      for entry in row: # go through each value (column) in the row
        cell = f"{chr(col_count)}{row_count}"
        if cell in value_locations.keys():
          values.update({value_locations[cell]:entry})
        col_count += 1
    return values


  # TODO: assume sheet is workbook's active sheet
  def update_value(self, value_name, value, sheet, template=None):
    if template == None:
      t_path = self._get_template_path(sheet)
      template = self._load_template(t_path)
    value_locations = self._get_template_value_locations(template)
    for k, v in value_locations.items():
      if v == value_name:
        sheet[k] = value

  
  def save_workbook(self):
    print(f"Saving workbook {self.file_path}")
    self.save(filename=self.file_path)


